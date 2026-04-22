from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from .models import ExpertAnalysisRun, StudentExpertData, CohortExpertData, SchoolSettings, HistoricalStudent, HistoricalGrade
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

@login_required
def expert_analysis_view(request):
    """
    Renders the Expert Analysis Dashboard (الخبراء).
    Requires 'director' or equivalent permission.
    """
    if not getattr(request.user, 'profile', None) or not request.user.profile.has_perm('access_analytics'):
        return render(request, 'students/unauthorized.html')

    # Fetch latest run data (defer new fields so page loads even if migration not yet applied)
    latest_run = ExpertAnalysisRun.objects.order_by('-run_date').first()
    expert_runs = list(ExpertAnalysisRun.objects.order_by('-run_date')[:15])
    try:
        cohort_sample = list(CohortExpertData.objects.defer('last_year_raw_avg').select_related('run').order_by('-run__run_date')[:20])
    except Exception:
        cohort_sample = list(CohortExpertData.objects.select_related('run').order_by('-run__run_date')[:20])
    student_expert_sample = list(StudentExpertData.objects.select_related('run', 'student').filter(student__isnull=False).order_by('-run__run_date')[:30])

    from .school_year_utils import get_current_school_year, get_prev_school_year, get_school_year_before_prev
    current_year = get_current_school_year()
    prev_year = get_prev_school_year(current_year)
    prev_year_before = get_school_year_before_prev(current_year)

    # إحصائيات "التاريخي" يجب أن تعتمد فقط على تلاميذ آخر فصل مستورد من السنة الحالية
    from .models import Grade
    term_order = {'الفصل الأول': 1, 'الفصل الثاني': 2, 'الفصل الثالث': 3}
    terms = list(Grade.objects.filter(academic_year=current_year).values_list('term', flat=True).distinct())
    terms = [t for t in terms if t]
    latest_term = max(terms, key=lambda t: term_order.get(t, 0)) if terms else None
    qs_ref = Grade.objects.filter(academic_year=current_year)
    if latest_term:
        qs_ref = qs_ref.filter(term=latest_term)
    current_ids = list(qs_ref.values_list('student__student_id_number', flat=True).distinct())
    current_ids = [x for x in current_ids if x]
    current_students_count = len(current_ids)
    # HistoricalStudent يحتوي سجلاً لكل (تلميذ، سنة تاريخية) لذلك نعدّ التلاميذ المميزين فقط
    current_students_with_archive_count = HistoricalStudent.objects.filter(student_id_number__in=current_ids).values('student_id_number').distinct().count() if current_ids else 0
    # الإبقاء على الاسم القديم للتوافق مع القالب الحالي
    historical_student_count = current_students_with_archive_count
    historical_grade_count = HistoricalGrade.objects.filter(student__student_id_number__in=current_ids).count() if current_ids else 0

    context = {
        'page_title': 'تحليل خبراء التربية',
        'has_data': latest_run is not None and latest_run.status == 'completed',
        'latest_run': latest_run,
        'current_year': current_year,
        'prev_year': prev_year,
        'prev_year_before': prev_year_before,
        'current_year_latest_term': latest_term,
        'expert_runs': expert_runs,
        'cohort_sample': cohort_sample,
        'student_expert_sample': student_expert_sample,
        'historical_student_count': historical_student_count,
        'current_students_count': current_students_count,
        'current_students_with_archive_count': current_students_with_archive_count,
        'historical_grade_count': historical_grade_count,
    }

    return render(request, 'students/expert_analysis.html', context)


@login_required
def expert_delete_run(request, run_id):
    """حذف تشغيلة تحليل الخبراء (والبيانات المرتبطة cascade)."""
    if not getattr(request.user, 'profile', None) or not request.user.profile.has_perm('access_analytics'):
        return JsonResponse({'status': 'error', 'message': 'غير مصرح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'طريقة غير صالحة'}, status=405)
    run = get_object_or_404(ExpertAnalysisRun, id=run_id)
    run.delete()
    messages.success(request, f'تم حذف التشغيلة {run_id}.')
    return redirect('expert_analysis_view')


@login_required
def expert_export_excel(request):
    """تصدير نتائج تحليل الخبراء إلى Excel (كاشف الأنماط + بيانات الأفواج)."""
    if not getattr(request.user, 'profile', None) or not request.user.profile.has_perm('access_analytics'):
        return HttpResponse('Unauthorized', status=403)
    latest_run = ExpertAnalysisRun.objects.filter(status='completed').order_by('-run_date').first()
    if not latest_run:
        messages.warning(request, 'لا توجد بيانات خبراء معالجة للتصدير.')
        return redirect('expert_analysis_view')
    level = request.GET.get('level', '')
    students_qs = StudentExpertData.objects.filter(run=latest_run).select_related('student')
    if level:
        students_qs = students_qs.filter(academic_year_level__icontains=level)
    cohort_qs = CohortExpertData.objects.filter(run=latest_run)
    if level:
        cohort_qs = cohort_qs.filter(academic_year_level__icontains=level)

    wb = openpyxl.Workbook()
    # ورقة كاشف الأنماط
    ws1 = wb.active
    ws1.title = "كاشف الأنماط"
    headers1 = ['التلميذ', 'القسم', 'المستوى', 'المعدل الحالي', 'المتوقع', 'الباقي', 'النمط', 'Z-Score', 'القيمة المضافة']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
        ws1.cell(row=1, column=col).font = Font(bold=True)
    for row, s in enumerate(students_qs, 2):
        ws1.cell(row=row, column=1, value=s.student.full_name if s.student else '')
        ws1.cell(row=row, column=2, value=s.class_name)
        ws1.cell(row=row, column=3, value=s.academic_year_level)
        ws1.cell(row=row, column=4, value=round(s.current_avg, 2) if s.current_avg is not None else '')
        ws1.cell(row=row, column=5, value=round(s.predicted_avg, 2) if s.predicted_avg is not None else '')
        ws1.cell(row=row, column=6, value=round(s.residual, 2) if s.residual is not None else '')
        ws1.cell(row=row, column=7, value=s.status_pattern or '')
        ws1.cell(row=row, column=8, value=round(s.z_score, 2) if s.z_score is not None else '')
        ws1.cell(row=row, column=9, value=round(s.net_value_added, 2) if s.net_value_added is not None else '')
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 14
    # ورقة الأفواج والمادة الحاكمة
    ws2 = wb.create_sheet("الأفواج والمادة الحاكمة")
    headers2 = ['المستوى', 'المادة الحاكمة', 'متوسط Z الحالي', 'متوسط Z السابق', 'تحليل أثر الفوج']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
        ws2.cell(row=1, column=col).font = Font(bold=True)
    for row, c in enumerate(cohort_qs, 2):
        ws2.cell(row=row, column=1, value=c.academic_year_level)
        ws2.cell(row=row, column=2, value=c.ruling_subject or '')
        ws2.cell(row=row, column=3, value=round(c.current_year_z_score_avg, 3) if c.current_year_z_score_avg is not None else '')
        ws2.cell(row=row, column=4, value=round(c.last_year_z_score_avg, 3) if c.last_year_z_score_avg is not None else '')
        ws2.cell(row=row, column=5, value=c.cohort_effect_analysis or '')
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"تحليل_خبراء_التربية_{latest_run.run_date.strftime('%Y-%m-%d_%H-%M')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
