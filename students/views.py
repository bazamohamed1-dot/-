from rest_framework import viewsets
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db.models import Count, F, Q
from django.utils import timezone
from .models import (
    Student,
    CanteenAttendance,
    CanteenDailySummary,
    LibraryLoan,
    SchoolSettings,
    ArchiveDocument,
    EmployeeProfile,
    SystemMessage,
    PendingUpdate,
    AttendanceRecord,
    Communication,
)
from .serializers import StudentSerializer, StudentListSerializer, CanteenAttendanceSerializer, LibraryLoanSerializer, SchoolSettingsSerializer, ArchiveDocumentSerializer, SystemMessageSerializer, PendingUpdateSerializer
from .utils import normalize_arabic
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse, FileResponse
from datetime import date, timedelta, datetime
from collections import defaultdict
import os
from io import BytesIO
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
import logging
import base64
import uuid

logger = logging.getLogger(__name__)

def service_worker(request):
    return render(request, 'sw.js', content_type='application/javascript')

def parse_smart_date(date_val):
    if not date_val: return date(1900, 1, 1)
    date_str = str(date_val).strip()
    if date_str.lower() in ['none', 'nan', '', 'invalid date']: return date(1900, 1, 1)

    # Already YYYY-MM-DD (standard JSON)
    if len(date_str) == 10 and date_str[4] == '-':
            try: return date.fromisoformat(date_str)
            except: pass

    # Handle Excel Serial Date
    if str(date_str).replace('.', '', 1).isdigit():
            try:
                val = float(date_str)
                if val > 10000:
                    return (datetime(1899, 12, 30) + timedelta(days=val)).date()
            except: pass

    # Clean time part if present
    date_str = date_str.split('T')[0].split(' ')[0]

    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y', '%Y.%m.%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return date(1900, 1, 1)

# --- Restore StudentViewSet ---
class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Allow basic filtering by 'academic_year' (level) and 'class_name' (class)
        # This supports server-side filtering for the management interface.
        queryset = Student.objects.all().order_by('id')

        level = self.request.query_params.get('academic_year') or self.request.query_params.get('level')
        class_name = self.request.query_params.get('class_name') or self.request.query_params.get('class')
        search = self.request.query_params.get('search') or self.request.query_params.get('q')

        if level:
            queryset = queryset.filter(academic_year=level)
        if class_name:
            # We want exact match for class_name since it's just a number now
            queryset = queryset.filter(class_name=class_name)
        if search:
            norm_search = normalize_arabic(search)
            q_obj = Q(student_id_number__icontains=search) | \
                    Q(first_name__icontains=search) | \
                    Q(last_name__icontains=search)

            # Add normalized search
            if norm_search != search:
                q_obj |= Q(first_name__icontains=norm_search) | \
                         Q(last_name__icontains=norm_search)

            # Common variations (Ha/Ta Marbuta, Alef)
            if 'ه' in search:
                 alt = search.replace('ه', 'ة')
                 q_obj |= Q(first_name__icontains=alt) | Q(last_name__icontains=alt)
            if 'ة' in search:
                 alt = search.replace('ة', 'ه')
                 q_obj |= Q(first_name__icontains=alt) | Q(last_name__icontains=alt)

            queryset = queryset.filter(q_obj)

        return queryset

    def get_serializer_class(self):
        # Use lightweight serializer for list view to prevent memory crash
        if self.action == 'list':
            return StudentListSerializer
        return StudentSerializer

    def create(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('student_add'):
             return Response({'error': 'Unauthorized'}, status=403)

        # If user is Director or Superuser, apply immediately
        if request.user.profile.role == 'director' or request.user.is_superuser:
            return super().create(request, *args, **kwargs)

        # For other users, create a Pending Update
        try:
            PendingUpdate.objects.create(
                user=request.user,
                model_name='Student',
                action='create',
                data=request.data,
                status='pending'
            )
            return Response({'message': 'تم إرسال التلميذ الجديد إلى المدير للموافقة', 'pending': True}, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('student_edit'):
             return Response({'error': 'Unauthorized'}, status=403)

        # If user is Director or Superuser, apply immediately
        if request.user.profile.role == 'director' or request.user.is_superuser:
            return super().update(request, *args, **kwargs)

        # For other users, create a Pending Update
        student = self.get_object()

        try:
            PendingUpdate.objects.create(
                user=request.user,
                model_name='Student',
                action='update',
                data={
                    'id': student.id,
                    **request.data
                },
                status='pending'
            )
            return Response({'message': 'تم إرسال التحديث إلى المدير للموافقة', 'pending': True}, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('student_delete'):
             return Response({'error': 'Unauthorized'}, status=403)

        # If user is Director or Superuser, apply immediately
        if request.user.profile.role == 'director' or request.user.is_superuser:
            return super().destroy(request, *args, **kwargs)

        student = self.get_object()

        try:
            PendingUpdate.objects.create(
                user=request.user,
                model_name='Student',
                action='delete',
                data={'id': student.id, 'name': str(student)},
                status='pending'
            )
            return Response({'message': 'تم إرسال طلب الحذف إلى المدير', 'pending': True}, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('student_delete'):
             return Response({'error': 'Unauthorized'}, status=403)

        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'No IDs provided'}, status=status.HTTP_400_BAD_REQUEST)

        # If user is Director or Superuser, apply immediately
        if request.user.profile.role == 'director' or request.user.is_superuser:
            try:
                Student.objects.filter(id__in=ids).delete()
                return Response({'message': f'Deleted {len(ids)} students'})
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Pending Update for Bulk Delete
        try:
            PendingUpdate.objects.create(
                user=request.user,
                model_name='Student',
                action='delete', # Using delete action, but data has list of IDs
                data={'ids': ids, 'is_bulk': True},
                status='pending'
            )
            return Response({'message': 'تم إرسال طلب الحذف الجماعي إلى المدير', 'pending': True}, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_all(self, request):
        # Allow export for management, import_data, or director
        has_perm = False
        if request.user.is_superuser:
            has_perm = True
        elif hasattr(request.user, 'profile'):
            if request.user.profile.role == 'director':
                has_perm = True
            elif request.user.profile.has_perm('import_data') or request.user.profile.has_perm('access_management'):
                has_perm = True

        if not has_perm:
             return Response({'error': 'Unauthorized'}, status=403)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Backup_Students.xlsx'

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Students")

        # Headers
        headers = [
            "رقم التعريف الوطني", "اللقب", "الاسم", "تاريخ الازدياد", "مكان الميلاد",
            "الجنس", "المستوى", "القسم", "نظام التمدرس", "رقم القيد",
            "تاريخ التسجيل", "تاريخ الخروج", "اسم الولي", "اسم الأم",
            "هاتف الولي", "العنوان", "مسار الصورة"
        ]
        ws.append(headers)

        # Stream Data
        students = Student.objects.all().order_by('academic_year', 'class_name', 'last_name')
        for s in students.iterator():
            ws.append([
                s.student_id_number, s.last_name, s.first_name,
                s.date_of_birth, s.place_of_birth, s.gender,
                s.academic_year, s.class_name, s.attendance_system,
                s.enrollment_number, s.enrollment_date, s.exit_date,
                s.guardian_name, s.mother_name, s.guardian_phone,
                s.address, s.photo.name if s.photo else ''
            ])

        wb.save(response)
        return response

class PendingUpdateViewSet(viewsets.ModelViewSet):
    queryset = PendingUpdate.objects.all().order_by('-timestamp')
    serializer_class = PendingUpdateSerializer
    permission_classes = [IsAuthenticated]

    def _is_director_or_superuser(self):
        if self.request.user.is_superuser:
            return True
        if hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'director':
            return True
        return False

    def get_queryset(self):
        # Allow operations on the endpoint regardless of 'status' for Director,
        # but default to pending
        if self._is_director_or_superuser():
            return PendingUpdate.objects.all()
        # Regular users see their own pending updates to reflect in UI
        return PendingUpdate.objects.filter(user=self.request.user, status='pending')

    def _apply_update(self, update):
        # Handle inconsistent data structure (nested 'data' key vs flat)
        data_payload = update.data.get('data', update.data)
        if isinstance(data_payload, str): # Handle stringified JSON
             import json
             try: data_payload = json.loads(data_payload)
             except: pass

        if update.model_name == 'Student':
            # Handle Photo: If it's a new Base64 string, it should be processed.
            # If it's an existing URL (string not starting with data:image), we REMOVE it from payload
            # so the serializer (partial=True) doesn't touch the existing photo.
            if 'photo' in data_payload:
                photo_val = data_payload['photo']
                if isinstance(photo_val, str):
                    if photo_val.startswith('data:image'):
                        # Valid Base64 - Keep it, it will overwrite existing.
                        pass
                    else:
                        # Existing URL or Path - Remove to prevent clearing/error.
                        del data_payload['photo']
                elif photo_val is None:
                    # Explicit null might mean "delete photo".
                    # However, in offline context, user might not have cleared it.
                    # Safety check: if user INTENDS to delete, we might need a flag.
                    # For now, let's assume null means "no change" unless we have a specific delete flag.
                    # But serializer allows null. Let's rely on frontend sending "" for no change or null for delete?
                    # Frontend (management.html) logic:
                    # if (photoVal && photoVal.startsWith('data:image')) { studentData.photo = photoVal; }
                    # So frontend DOES NOT send photo if it's not changed (it sends nothing).
                    # If data_payload has 'photo': null, it might be an explicit clear.
                    pass

            # Use serializer to validate and save
            if update.action == 'create':
                serializer = StudentSerializer(data=data_payload)
                if serializer.is_valid():
                    serializer.save()
                else:
                    logger.error(f"Create Student Failed: {serializer.errors}")
                    raise ValueError(f"Invalid data: {serializer.errors}")

            elif update.action == 'update':
                obj_id = data_payload.get('id')
                if obj_id:
                    try:
                        student = Student.objects.get(id=obj_id)
                        serializer = StudentSerializer(student, data=data_payload, partial=True)
                        if serializer.is_valid():
                            serializer.save()
                        else:
                            logger.error(f"Update Student Failed: {serializer.errors}")
                            raise ValueError(f"Invalid data: {serializer.errors}")
                    except Student.DoesNotExist:
                        logger.warning(f"Student {obj_id} not found for update")
                        pass

            elif update.action == 'delete':
                if data_payload.get('is_bulk'):
                    ids = data_payload.get('ids', [])
                    Student.objects.filter(id__in=ids).delete()
                else:
                    obj_id = data_payload.get('id')
                    if obj_id:
                        Student.objects.filter(id=obj_id).delete()

        elif update.model_name == 'CanteenAttendance':
            # Logic for canteen offline sync
            barcode = data_payload.get('barcode') or data_payload.get('student_id')
            student = None
            if barcode:
                student = Student.objects.filter(student_id_number=barcode).first()
                if not student and str(barcode).isdigit():
                        student = Student.objects.filter(id=barcode).first()

            if student:
                if not CanteenAttendance.objects.filter(student=student, date=date.today()).exists():
                    reg = data_payload.get('registration_method')
                    if reg not in (CanteenAttendance.REG_SCAN, CanteenAttendance.REG_MANUAL):
                        reg = (
                            CanteenAttendance.REG_MANUAL
                            if data_payload.get('student_id') and not data_payload.get('barcode')
                            else CanteenAttendance.REG_SCAN
                        )
                    CanteenAttendance.objects.create(
                        student=student,
                        date=date.today(),
                        registration_method=reg,
                    )

        update.delete()

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not self._is_director_or_superuser():
             return Response({'error': 'Unauthorized'}, status=403)

        try:
            # Bypass get_object() which relies on get_queryset() to ensure it finds it
            # even if the queryset filtering might have an edge case.
            update = PendingUpdate.objects.get(pk=pk)
        except PendingUpdate.DoesNotExist:
            return Response({'error': 'التحديث غير موجود أو تمت الموافقة عليه مسبقاً.'}, status=404)

        try:
            # Create notification
            if update.user:
                SystemMessage.objects.create(
                    recipient=update.user,
                    message=f"تم قبول التحديث الخاص بـ {update.model_name} (ID: {update.data.get('id')})",
                    active=True
                )

            self._apply_update(update)
            return Response({'message': 'Approved'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['post'])
    def approve_all(self, request):
        if not self._is_director_or_superuser():
             return Response({'error': 'Unauthorized'}, status=403)

        updates = self.get_queryset()
        count = 0
        errors = []
        for update in updates:
             try:
                if update.user:
                    SystemMessage.objects.create(
                        recipient=update.user,
                        message=f"تم قبول التحديث الخاص بـ {update.model_name} (ID: {update.data.get('id')})",
                        active=True
                    )
                self._apply_update(update)
                count += 1
             except Exception as e:
                errors.append(f"ID {update.id}: {str(e)}")

        return Response({'message': f'Approved {count}', 'errors': errors})

    @action(detail=False, methods=['post'])
    def reject_all(self, request):
        if not self._is_director_or_superuser():
             return Response({'error': 'Unauthorized'}, status=403)

        updates = self.get_queryset()
        for update in updates:
            if update.user:
                SystemMessage.objects.create(
                    recipient=update.user,
                    message=f"تم رفض التحديث الخاص بـ {update.model_name} (ID: {update.data.get('id')})",
                    active=True
                )

        updates.delete()
        return Response({'message': 'Rejected All'})

    @action(detail=False, methods=['get'])
    def count(self, request):
        if not self._is_director_or_superuser():
             return Response({'count': 0})
        count = self.get_queryset().count()
        return Response({'count': count})

    # Sync Endpoint for Offline Manager
    # This receives a list of requests from the frontend and converts them to PendingUpdates
    @action(detail=False, methods=['post'], url_path='sync')
    def sync_offline(self, request):
        items = request.data
        if not isinstance(items, list):
            items = [items]

        count = 0
        import json
        for item in items:
            url = item.get('url', '')
            method = item.get('method', 'POST')
            payload = item.get('body', {})

            # Ensure payload is a dictionary, not a stringified JSON from FormData
            if isinstance(payload, str):
                try:
                    payload = json.loads(payload)
                except:
                    pass # Keep as string if not valid JSON

            # Map URL to Model
            model_name = 'Unknown'
            action_type = 'create'

            if 'students' in url:
                model_name = 'Student'
                if method == 'PUT' or method == 'PATCH': action_type = 'update'
                elif method == 'DELETE': action_type = 'delete'
                elif 'bulk_delete' in url: action_type = 'delete'
            elif 'scan_card' in url or 'manual_attendance' in url:
                model_name = 'CanteenAttendance'
            elif 'loan' in url:
                model_name = 'LibraryLoan'

            # Create Pending Update directly
            # Note: We trust the authenticated user here.
            PendingUpdate.objects.create(
                user=request.user,
                model_name=model_name,
                action=action_type,
                data=payload,
                status='pending'
            )
            count += 1

        return Response({'message': f'Synced {count} items'})


# --- Lightweight JSON Import API ---
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_students_json(request):
    """
    Receives a JSON list of students and performs bulk create/update.
    This avoids server-side file parsing memory overhead.
    """
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('import_data'):
         return Response({'error': 'Unauthorized'}, status=403)

    data = request.data
    students_list = data.get('students', [])
    update_existing = data.get('update_existing', False)

    if not students_list:
        return Response({'error': 'No data provided'}, status=400)

    created_count = 0
    updated_count = 0
    errors = []

    # Map existing students for fast lookup
    existing_map = {s.student_id_number: s for s in Student.objects.all()}

    to_create = []
    to_update = []

    for item in students_list:
        try:
            sid = str(item.get('student_id_number')).strip()
            if not sid: continue

            # Robust Date Parsing
            dob = parse_smart_date(item.get('date_of_birth'))
            enroll_date = parse_smart_date(item.get('enrollment_date'))
            if enroll_date == date(1900, 1, 1): enroll_date = date.today()

            student_data = {
                'student_id_number': sid,
                'last_name': item.get('last_name', ''),
                'first_name': item.get('first_name', ''),
                'gender': item.get('gender', ''),
                'date_of_birth': dob,
                'place_of_birth': item.get('place_of_birth', ''),
                'academic_year': item.get('academic_year', ''),
                'class_name': item.get('class_name', ''),
                'attendance_system': item.get('attendance_system', ''),
                'enrollment_number': item.get('enrollment_number', ''),
                'enrollment_date': enroll_date,
                'guardian_name': item.get('guardian_name', ''),
                'mother_name': item.get('mother_name', ''),
                'address': item.get('address', ''),
                'guardian_phone': item.get('guardian_phone', ''),
            }

            if sid in existing_map:
                if update_existing:
                    s = existing_map[sid]
                    for key, val in student_data.items():
                        if key != 'student_id_number': # Don't update PK
                            setattr(s, key, val)
                    to_update.append(s)
            else:
                to_create.append(Student(**student_data))

        except Exception as e:
            errors.append(f"Row error: {str(e)}")

    try:
        if to_create:
            Student.objects.bulk_create(to_create)
            created_count = len(to_create)

        if to_update:
            Student.objects.bulk_update(to_update, [
                'last_name', 'first_name', 'gender', 'date_of_birth', 'place_of_birth',
                'academic_year', 'class_name', 'attendance_system', 'enrollment_number',
                'enrollment_date', 'guardian_name', 'mother_name', 'address', 'guardian_phone'
            ])
            updated_count = len(to_update)

        return Response({
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)

class ArchiveDocumentViewSet(viewsets.ModelViewSet):
    queryset = ArchiveDocument.objects.all().order_by('-entry_date')
    serializer_class = ArchiveDocumentSerializer

    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'Archive_Export.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أرشيف المؤسسة"

        headers = ["الرقم", "المصلحة", "الملف/السجل", "الوثيقة", "الرمز", "تاريخ الازدياد", "تاريخ الدخول", "تاريخ الخروج المؤقت", "تاريخ الحذف", "ملاحظات"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        docs = self.filter_queryset(self.get_queryset())
        for doc in docs:
            ws.append([
                doc.reference_number,
                doc.service,
                doc.file_type,
                doc.document_type,
                doc.symbol or '',
                str(doc.student_dob) if doc.student_dob else '',
                str(doc.entry_date),
                str(doc.temp_exit_date) if doc.temp_exit_date else '',
                str(doc.elimination_date) if doc.elimination_date else '',
                doc.notes or ''
            ])

        wb.save(file_path)
        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Archive_{date.today()}.xlsx"'
        return response

# --- Library Views ---

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def scan_library_card(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('library_scan'):
        return Response({'error': 'Unauthorized'}, status=403)
    try:
        barcode = request.data.get('barcode')
        if not barcode:
            return Response({'error': 'No barcode provided'}, status=status.HTTP_400_BAD_REQUEST)

        barcode = str(barcode).strip()

        # Robust lookup
        try:
            student = Student.objects.get(student_id_number=barcode)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found', 'code': 'NOT_FOUND'}, status=status.HTTP_404_NOT_FOUND)
        except Student.MultipleObjectsReturned:
            # Fallback: take the first one or error out. Usually shouldn't happen with ID.
            student = Student.objects.filter(student_id_number=barcode).first()
            if not student:
                 return Response({'error': 'Student not found', 'code': 'NOT_FOUND'}, status=status.HTTP_404_NOT_FOUND)

        # Get active loans
        active_loans = LibraryLoan.objects.filter(student=student, is_returned=False)

        # Check limit
        settings_obj = SchoolSettings.objects.first()
        limit = settings_obj.loan_limit if settings_obj else 2
        limit_reached = active_loans.count() >= limit

        return Response({
            'student': StudentSerializer(student).data,
            'active_loans': LibraryLoanSerializer(active_loans, many=True).data,
            'limit_reached': limit_reached,
            'loan_limit': limit
        })
    except Exception as e:
        logger.error(f"Library Scan Error: {e}")
        return Response({'error': f'Server Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_loan(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('library_loan'):
        return Response({'error': 'Unauthorized'}, status=403)
    student_id = request.data.get('student_id')
    book_title = request.data.get('book_title')
    loan_date_str = request.data.get('loan_date') # Allow manual override

    if not student_id or not book_title:
        return Response({'error': 'Missing data'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

    # Check loan limit
    active_loans_count = LibraryLoan.objects.filter(student=student, is_returned=False).count()
    settings_obj = SchoolSettings.objects.first()
    limit = 2
    if settings_obj:
        # Check specific level limit
        level = student.academic_year
        if level and settings_obj.loan_limits_by_level and level in settings_obj.loan_limits_by_level:
            try:
                limit = int(settings_obj.loan_limits_by_level[level])
            except:
                limit = settings_obj.loan_limit
        else:
            limit = settings_obj.loan_limit

    if active_loans_count >= limit:
        return Response({'error': f'لا يمكن استعارة أكثر من {limit} كتب'}, status=status.HTTP_400_BAD_REQUEST)

    if loan_date_str:
        try:
            loan_date = date.fromisoformat(loan_date_str)
        except ValueError:
            return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        loan_date = date.today()

    expected_return = loan_date + timedelta(days=15)

    loan = LibraryLoan.objects.create(
        student=student,
        book_title=book_title,
        loan_date=loan_date,
        expected_return_date=expected_return
    )

    return Response(LibraryLoanSerializer(loan).data, status=status.HTTP_201_CREATED)

@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def get_readers(request):
    if request.method == 'DELETE':
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('library_readers_list'):
             return Response({'error': 'Unauthorized'}, status=403)

        count = LibraryLoan.objects.all().delete()[0]
        return Response({'message': f'Deleted {count} records'})

    # Point 12: Include loan date and book title
    # Since a student can have multiple loans, returning "distinct student" hides details.
    # The user wants "When clicking reader list, show loan date and book".
    # This implies listing LOANS, not just unique students.
    loans = LibraryLoan.objects.select_related('student').order_by('-loan_date', '-loan_time')
    data = []
    for loan in loans:
        s = loan.student
        data.append({
            'student_id_number': s.student_id_number,
            'first_name': s.first_name,
            'last_name': s.last_name,
            'class_name': s.class_name,
            'book_title': loan.book_title,
            'loan_date': loan.loan_date,
            'loan_time': loan.loan_time.strftime("%H:%M:%S") if loan.loan_time else ""
        })
    return Response(data)

@csrf_exempt
@api_view(['GET', 'POST'])
def school_settings(request):
    settings_obj = SchoolSettings.objects.first()

    if request.method == 'GET':
        if settings_obj:
            return Response(SchoolSettingsSerializer(settings_obj).data)
        return Response({})

    elif request.method == 'POST':
        if settings_obj:
            serializer = SchoolSettingsSerializer(settings_obj, data=request.data, partial=True)
        else:
            serializer = SchoolSettingsSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        print(f"Settings Save Error: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def return_book(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('library_return'):
        return Response({'error': 'Unauthorized'}, status=403)

    loan_id = request.data.get('loan_id')
    try:
        loan = LibraryLoan.objects.get(id=loan_id)
    except LibraryLoan.DoesNotExist:
        return Response({'error': 'Loan not found'}, status=status.HTTP_404_NOT_FOUND)

    loan.is_returned = True
    loan.actual_return_date = date.today()
    loan.save()

    return Response({'message': 'Book returned successfully'})

@api_view(['GET'])
def library_stats(request):
    today = date.today()

    # Point 10: Stats Intervals (Daily, Weekly, Monthly, Quarterly, Yearly)
    # Using loan_date as the metric
    daily = LibraryLoan.objects.filter(loan_date=today).count()

    week_start = today - timedelta(days=today.weekday()) # Start of week (Monday?) or just last 7 days? User said "Weekly". Let's do last 7 days.
    # Actually "Weekly" usually means "This Week". But let's stick to simple "Last 7 days" or "This week" depending on interpretation.
    # Let's use "This Week" (since start of week) or rolling 7 days. Rolling 7 is safer.
    weekly = LibraryLoan.objects.filter(loan_date__gte=today - timedelta(days=7)).count()
    monthly = LibraryLoan.objects.filter(loan_date__gte=today - timedelta(days=30)).count()
    quarterly = LibraryLoan.objects.filter(loan_date__gte=today - timedelta(days=90)).count()
    yearly = LibraryLoan.objects.filter(loan_date__gte=today - timedelta(days=365)).count()

    total_students = Student.objects.count()
    borrowers_count = LibraryLoan.objects.values('student').distinct().count()

    percentage = 0
    if total_students > 0:
        percentage = round((borrowers_count / total_students) * 100, 1)

    # Overdue loans: Not returned AND expected_return < today
    overdue_loans = LibraryLoan.objects.filter(
        is_returned=False,
        expected_return_date__lt=today
    ).select_related('student')

    overdue_list = []
    for loan in overdue_loans:
        overdue_list.append({
            'student_name': f"{loan.student.last_name} {loan.student.first_name}",
            'student_class': loan.student.class_name,
            'book_title': loan.book_title,
            'loan_date': loan.loan_date,
            'expected_return': loan.expected_return_date,
            'days_overdue': (today - loan.expected_return_date).days
        })

    # Distribution stats
    class_dist = LibraryLoan.objects.values('student__class_name').annotate(count=Count('student', distinct=True)).order_by('student__class_name')
    level_dist = LibraryLoan.objects.values('student__academic_year').annotate(count=Count('student', distinct=True)).order_by('student__academic_year')

    return Response({
        'borrowers_count': borrowers_count,
        'borrowers_percentage': percentage,
        'overdue_loans': overdue_list,
        'stats': {
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'quarterly': quarterly,
            'yearly': yearly
        },
        'class_distribution': list(class_dist),
        'level_distribution': list(level_dist)
    })

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def scan_card(request):
    try:
        # Check permission
        if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('canteen_scan'):
             return Response({'error': 'Unauthorized'}, status=403)

        barcode = request.data.get('barcode')
        logger.info(f"DEBUG: Received barcode: {barcode}")

        if not barcode:
            return Response({'error': 'No barcode provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Clean the barcode (remove spaces, etc)
        barcode = str(barcode).strip()

        # Robust lookup
        student = None
        employee = None
        is_employee = False

        try:
            student = Student.objects.get(student_id_number=barcode)
        except Student.DoesNotExist:
            try:
                employee = Employee.objects.get(employee_code=barcode)
                is_employee = True
            except Employee.DoesNotExist:
                return Response({'error': 'Not found', 'code': 'NOT_FOUND'}, status=status.HTTP_404_NOT_FOUND)
            except Employee.MultipleObjectsReturned:
                employee = Employee.objects.filter(employee_code=barcode).first()
                if not employee:
                    return Response({'error': 'Not found', 'code': 'NOT_FOUND'}, status=status.HTTP_404_NOT_FOUND)
                is_employee = True
        except Student.MultipleObjectsReturned:
            student = Student.objects.filter(student_id_number=barcode).first()
            if not student:
                return Response({'error': 'Not found', 'code': 'NOT_FOUND'}, status=status.HTTP_404_NOT_FOUND)

        # Time Restriction Logic (Dynamic)
        settings_obj = SchoolSettings.objects.first()
        close_time = settings_obj.canteen_close_time if settings_obj else time(13, 15)
        open_time = settings_obj.canteen_open_time if settings_obj else time(12, 0)

        # Parse days (comma separated string "0,1,2")
        allowed_days = [0, 2, 3, 6] # Default Sun, Mon, Wed, Thu
        if settings_obj and settings_obj.canteen_days:
            try:
                allowed_days = [int(d) for d in settings_obj.canteen_days.split(',') if d.strip().isdigit()]
            except: pass

        now = timezone.localtime()
        current_time = now.time()
        current_weekday = now.weekday() # Mon=0, Sun=6

        # Schedule Check (Strict for EVERYONE, including Director)
        # Check Day
        if current_weekday not in allowed_days:
             return Response({
                 'error': 'المطعم مغلق اليوم',
                 'code': 'CLOSED_DAY',
                 'student': StudentSerializer(student).data if not is_employee else None
             }, status=status.HTTP_403_FORBIDDEN)

        # Check Time
        if current_time < open_time:
             return Response({
                 'error': f'المطعم يفتح على الساعة {open_time.strftime("%H:%M")}',
                 'code': 'NOT_OPEN_YET',
                 'student': StudentSerializer(student).data if not is_employee else None
             }, status=status.HTTP_403_FORBIDDEN)

        if current_time > close_time:
             today = date.today()
             if is_employee:
                 attended = CanteenAttendance.objects.filter(employee=employee, date=today).exists()
                 student_data = None
             else:
                 attended = CanteenAttendance.objects.filter(student=student, date=today).exists()
                 student_data = StudentSerializer(student).data

             if attended:
                 return Response({
                     'error': 'انتهى الوقت (أكل مسبقاً)',
                     'code': 'LATE_ATE',
                     'student': student_data
                 }, status=status.HTTP_403_FORBIDDEN)
             else:
                 return Response({
                     'error': 'انتهى الوقت (لم يأكل)',
                     'code': 'LATE_NOT_ATE',
                     'student': student_data
                 }, status=status.HTTP_403_FORBIDDEN)

        today = date.today()
        if is_employee:
            if CanteenAttendance.objects.filter(employee=employee, date=today).exists():
                return Response({'error': 'Already took the meal', 'code': 'ALREADY_ATE'}, status=status.HTTP_400_BAD_REQUEST)

            attendance = CanteenAttendance.objects.create(
                employee=employee,
                date=today,
                registration_method=CanteenAttendance.REG_SCAN,
            )
            return Response({
                'message': 'Attendance recorded',
                'is_employee': True,
                'employee_name': f"{employee.first_name} {employee.last_name}",
                'employee_rank': employee.get_rank_display(),
                'attendance': CanteenAttendanceSerializer(attendance).data
            }, status=status.HTTP_201_CREATED)

        else:
            # Check if Half-Board
            if student.attendance_system != 'نصف داخلي':
                return Response({
                    'error': 'Student is not Half-Board',
                    'student': StudentSerializer(student).data,
                    'code': 'NOT_HALF_BOARD'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if already attended today
            if CanteenAttendance.objects.filter(student=student, date=today).exists():
                return Response({
                    'error': 'Student already took the meal',
                    'student': StudentSerializer(student).data,
                    'code': 'ALREADY_ATE'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Record attendance
            attendance = CanteenAttendance.objects.create(
                student=student,
                date=today,
                registration_method=CanteenAttendance.REG_SCAN,
            )
            return Response({
                'message': 'Attendance recorded',
                'student': StudentSerializer(student).data,
                'attendance': CanteenAttendanceSerializer(attendance).data
            }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Canteen Scan Error: {e}")
        return Response({'error': f'Server Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_canteen_stats(request):
    today = date.today()
    total_half_board = Student.objects.filter(attendance_system='نصف داخلي').count()
    present_count = CanteenAttendance.objects.filter(date=today).count()

    # User requirement: If no one registered yet (or holiday), don't mark everyone as absent.
    if present_count == 0:
        absent_count = 0
    else:
        absent_count = total_half_board - present_count

    return Response({
        'total_half_board': total_half_board,
        'present_count': present_count,
        'absent_count': max(0, absent_count)
    })

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manual_attendance(request):
    # Check permission
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('canteen_manual'):
        return Response({'error': 'Unauthorized'}, status=403)

    # Time Restriction Logic (Dynamic)
    settings_obj = SchoolSettings.objects.first()
    close_time = settings_obj.canteen_close_time if settings_obj else time(13, 15)
    open_time = settings_obj.canteen_open_time if settings_obj else time(12, 0)

    # Parse days
    allowed_days = [0, 2, 3, 6]
    if settings_obj and settings_obj.canteen_days:
        try:
            allowed_days = [int(d) for d in settings_obj.canteen_days.split(',') if d.strip().isdigit()]
        except: pass

    now = timezone.localtime()
    current_time = now.time()
    current_weekday = now.weekday()

    # Enforce Schedule for EVERYONE
    if current_weekday not in allowed_days:
        return Response({'error': 'المطعم مغلق اليوم'}, status=status.HTTP_403_FORBIDDEN)

    if current_time < open_time:
        return Response({'error': f'المطعم يفتح على الساعة {open_time.strftime("%H:%M")}'}, status=status.HTTP_403_FORBIDDEN)

    if current_time > close_time:
        return Response({'error': f'انتهى وقت الإطعام ({close_time.strftime("%H:%M")})'}, status=status.HTTP_403_FORBIDDEN)

    student_id = request.data.get('student_id')
    today = date.today()

    student = None
    employee = None
    is_employee = False

    # Can search by internal ID or student_id_number/employee_code
    try:
        student = Student.objects.get(id=student_id)
    except (Student.DoesNotExist, ValueError):
        try:
            student = Student.objects.get(student_id_number=student_id)
        except Student.DoesNotExist:
            try:
                employee = Employee.objects.get(employee_code=student_id)
                is_employee = True
            except Employee.DoesNotExist:
                return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    if is_employee:
        if CanteenAttendance.objects.filter(employee=employee, date=today).exists():
            return Response({'error': 'Already recorded'}, status=status.HTTP_400_BAD_REQUEST)

        CanteenAttendance.objects.create(
            employee=employee,
            date=today,
            registration_method=CanteenAttendance.REG_MANUAL,
        )
        return Response({
            'message': 'Success',
            'is_employee': True,
            'employee_name': f"{employee.first_name} {employee.last_name}",
            'employee_rank': employee.get_rank_display()
        }, status=status.HTTP_201_CREATED)

    else:
        if student.attendance_system != 'نصف داخلي':
            return Response({'error': 'Student is not Half-Board'}, status=status.HTTP_400_BAD_REQUEST)

        if CanteenAttendance.objects.filter(student=student, date=today).exists():
            return Response({'error': 'Student already recorded'}, status=status.HTTP_400_BAD_REQUEST)

        CanteenAttendance.objects.create(
            student=student,
            date=today,
            registration_method=CanteenAttendance.REG_MANUAL,
        )
        return Response(
            {'message': 'Success', 'student': StudentSerializer(student).data},
            status=status.HTTP_201_CREATED
        )

@csrf_exempt
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_attendance(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'director':
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

    try:
        # Check for list of IDs or single ID in query params or body
        # Body: {"ids": [1, 2]} or {"id": 1}
        # Query: ?id=1

        ids = request.data.get('ids')
        single_id = request.data.get('id') or request.query_params.get('id')

        if ids:
            # Bulk delete by student IDs for today, or by attendance IDs?
            # Usually easiest to delete by Attendance ID if the frontend has it.
            # But the frontend might only have Student ID.
            # Let's assume Student IDs for robust "Remove Student from Today's List"

            # Actually, deleting by Attendance PK is safer if we have it.
            # If the input is student IDs, we must filter by today + student.

            # Let's support both: if 'type' param says 'student', delete by student. Else by PK.
            # Default: Attendance ID (PK).

            count, _ = CanteenAttendance.objects.filter(id__in=ids).delete()
            return Response({'message': f'تم حذف {count} سجلات'})

        elif single_id:
            # Check if we are deleting by student_id or attendance_id
            # Let's look up the attendance record first
            attendance = CanteenAttendance.objects.filter(id=single_id).first()
            if attendance:
                attendance.delete()
                return Response({'message': 'تم حذف السجل'})

            # Fallback: maybe it's a student ID and we want to remove today's entry?
            # Ideally the frontend sends the correct ID.
            # Let's enforce sending Attendance ID for precision.
            return Response({'error': 'السجل غير موجود'}, status=status.HTTP_404_NOT_FOUND)

        # Clear All for Today
        elif request.data.get('clear_all') == True:
            count, _ = CanteenAttendance.objects.filter(date=date.today()).delete()
            return Response({'message': f'تم مسح {count} سجلات لليوم'})

        return Response({'error': 'Missing parameters'}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_attendance_lists(request):
    today = date.today()
    present_attendances = CanteenAttendance.objects.filter(date=today).select_related('student')

    present_data = []
    present_ids = []
    method_labels = {
        CanteenAttendance.REG_SCAN: 'مسح البطاقة',
        CanteenAttendance.REG_MANUAL: 'إدخال يدوي',
    }
    for att in present_attendances:
        s_data = StudentSerializer(att.student).data
        s_data['attendance_time'] = att.time.strftime("%H:%M:%S")
        s_data['attendance_id'] = att.id # Include primary key for precise deletion
        s_data['registration_method'] = att.registration_method
        s_data['registration_method_label'] = method_labels.get(
            att.registration_method, att.registration_method
        )
        present_data.append(s_data)
        present_ids.append(att.student.id)

    absent_students = Student.objects.filter(attendance_system='نصف داخلي').exclude(id__in=present_ids)

    return Response({
        'present': present_data,
        'absent': StudentSerializer(absent_students, many=True).data
    })

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_canteen_sheet(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('canteen_export'):
        return Response({'error': 'Unauthorized'}, status=403)

    today = date.today()
    date_str = today.strftime("%Y-%m-%d")

    qs = CanteenAttendance.objects.select_related('student').order_by('date', 'student__class_name', 'student__last_name')
    if not qs.exists():
        return Response({'message': 'لا يوجد سجلات حضور لتصديرها'}, status=status.HTTP_200_OK)

    by_date = defaultdict(list)
    for att in qs.iterator(chunk_size=1000):
        by_date[att.date].append(att)

    method_labels = {
        CanteenAttendance.REG_SCAN: 'مسح البطاقة',
        CanteenAttendance.REG_MANUAL: 'إدخال يدوي',
    }

    wb = openpyxl.Workbook()
    first_sheet = True
    for d in sorted(by_date.keys()):
        title = d.strftime("%Y-%m-%d")[:31]
        if first_sheet:
            ws = wb.active
            ws.title = title
            first_sheet = False
        else:
            ws = wb.create_sheet(title=title)
        ws.append(
            [
                "التاريخ",
                "التوقيت",
                "طريقة التسجيل",
                "رقم التعريف",
                "الاسم",
                "اللقب",
                "القسم",
                "الحالة",
            ]
        )
        for att in by_date[d]:
            s = att.student
            label = method_labels.get(att.registration_method, att.registration_method or '')
            ws.append(
                [
                    str(att.date),
                    att.time.strftime("%H:%M:%S"),
                    label,
                    s.student_id_number,
                    s.first_name,
                    s.last_name,
                    s.class_name,
                    "حاضر",
                ]
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Canteen_Attendance_ByDay_{date_str}.xlsx"'
    return response


def _canteen_meals_map():
    settings_obj = SchoolSettings.objects.first()
    if not settings_obj or not settings_obj.canteen_meals_by_date:
        return {}
    data = settings_obj.canteen_meals_by_date
    return data if isinstance(data, dict) else {}


def _parse_iso_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


@csrf_exempt
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def canteen_meal_plans(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('access_canteen'):
        return Response({'error': 'Unauthorized'}, status=403)

    settings_obj = SchoolSettings.objects.first()
    if not settings_obj:
        return Response({'error': 'لا توجد إعدادات مؤسسة'}, status=400)

    if request.method == 'GET':
        return Response({'meals': _canteen_meals_map()})

    # POST — دمج مفاتيح التواريخ المرسلة، أو استبدال كامل عند replace=true
    incoming = request.data.get('meals')
    if not isinstance(incoming, dict):
        return Response({'error': 'يتوقع حقل meals ككائن (تاريخ -> نص)'}, status=400)

    replace_all = bool(request.data.get('replace'))

    def _normalize_meals_dict(raw):
        out = {}
        for k, v in raw.items():
            key = str(k).strip()[:10]
            if not key:
                continue
            val = (v or '').strip() if isinstance(v, str) else str(v or '')
            if val != '':
                out[key] = val
        return out

    if replace_all:
        current = _normalize_meals_dict(incoming)
    else:
        current = dict(_canteen_meals_map())
        for k, v in incoming.items():
            key = str(k).strip()[:10]
            if not key:
                continue
            val = (v or '').strip() if isinstance(v, str) else str(v or '')
            if val == '':
                current.pop(key, None)
            else:
                current[key] = val

    settings_obj.canteen_meals_by_date = current
    settings_obj.save(update_fields=['canteen_meals_by_date'])
    return Response({'ok': True, 'meals': current})


@csrf_exempt
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def canteen_daily_summary(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('access_canteen'):
        return Response({'error': 'Unauthorized'}, status=403)

    d = _parse_iso_date(request.query_params.get('date')) or date.today()

    if request.method == 'GET':
        meals = _canteen_meals_map()
        meal_today = meals.get(d.isoformat(), '') or meals.get(str(d), '')
        student_live = CanteenAttendance.objects.filter(date=d).count()
        row = CanteenDailySummary.objects.filter(date=d).first()
        if row:
            return Response(
                {
                    'date': d.isoformat(),
                    'meal_description': row.meal_description,
                    'student_count': row.student_count,
                    'supervisors_count': row.supervisors_count,
                    'staff_count': row.staff_count,
                    'teachers_count': row.teachers_count,
                    'workers_count': row.workers_count,
                    'guests_count': row.guests_count,
                    'notes': row.notes,
                    'total': row.total_beneficiaries,
                    'saved': True,
                    'meal_plan_text': meal_today,
                    'student_count_live': student_live,
                }
            )
        return Response(
            {
                'date': d.isoformat(),
                'meal_description': meal_today,
                'student_count': student_live,
                'supervisors_count': 0,
                'staff_count': 0,
                'teachers_count': 0,
                'workers_count': 0,
                'guests_count': 0,
                'notes': '',
                'total': student_live,
                'saved': False,
                'meal_plan_text': meal_today,
                'student_count_live': student_live,
            }
        )

    # POST — حفظ / تحديث الملخص
    payload = request.data
    meal_desc = (payload.get('meal_description') or '').strip()
    try:
        sc = int(payload.get('student_count', 0))
    except (TypeError, ValueError):
        sc = 0
    sc = max(0, sc)

    def _pi(name):
        try:
            return max(0, int(payload.get(name, 0)))
        except (TypeError, ValueError):
            return 0

    obj, _ = CanteenDailySummary.objects.update_or_create(
        date=d,
        defaults={
            'meal_description': meal_desc,
            'student_count': sc,
            'supervisors_count': _pi('supervisors_count'),
            'staff_count': _pi('staff_count'),
            'teachers_count': _pi('teachers_count'),
            'workers_count': _pi('workers_count'),
            'guests_count': _pi('guests_count'),
            'notes': (payload.get('notes') or '').strip(),
        },
    )
    return Response(
        {
            'ok': True,
            'date': d.isoformat(),
            'total': obj.total_beneficiaries,
            'saved': True,
        },
        status=status.HTTP_200_OK,
    )


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_canteen_stats_excel(request):
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('canteen_export'):
        return Response({'error': 'Unauthorized'}, status=403)

    rows = CanteenDailySummary.objects.order_by('date')
    if not rows.exists():
        return Response({'message': 'لا توجد بطاقات إحصائية محفوظة للتصدير'}, status=status.HTTP_200_OK)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("إحصائيات_المطعم_التراكمية")
    ws.append(
        [
            "التاريخ",
            "مكونات الوجبة",
            "التلاميذ",
            "المشرفون المرافقون",
            "الموظفون",
            "الأساتذة",
            "العمال",
            "الضيوف",
            "المجموع",
            "الملاحظات",
        ]
    )
    for r in rows.iterator():
        ws.append(
            [
                r.date.isoformat(),
                r.meal_description,
                r.student_count,
                r.supervisors_count,
                r.staff_count,
                r.teachers_count,
                r.workers_count,
                r.guests_count,
                r.total_beneficiaries,
                r.notes,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    today_str = date.today().strftime("%Y-%m-%d")
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Canteen_Stats_Cumulative_{today_str}.xlsx"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_filters(request):
    """
    Returns distinct Academic Years and Classes for dynamic dropdowns.
    Uses robust query to ensure all variations are captured.
    Now grouped primarily by class_code to avoid duplications.
    """
    import re

    # Exclude empty or null values
    levels_raw = list(Student.objects.exclude(academic_year__isnull=True).exclude(academic_year__exact='').values_list('academic_year', flat=True).distinct())

    classes_raw = list(Student.objects.exclude(class_name__isnull=True).exclude(class_name__exact='').values_list('class_name', flat=True).distinct())

    derived_levels = set(levels_raw)

    # Sort logically
    def custom_sort(item):
        if not item: return (999, item)
        # Extract leading number for logical sorting (e.g., '1', '2', '1AM', '2AM')
        match = re.search(r'\d+', item)
        if match:
            return (int(match.group()), item)
        return (999, item) # Put non-numbered items at the end

    levels = sorted(list(derived_levels), key=custom_sort)

    clean_classes = set()
    for c in classes_raw:
        clean_classes.add(str(c).strip())

    classes = sorted(list(clean_classes), key=custom_sort)

    # Build a mapping of level -> available classes
    level_class_map = {}
    for level in levels:
        level_classes_names = list(set(
            str(c).strip() for c in
            Student.objects.filter(academic_year=level)
            .exclude(class_name__isnull=True)
            .exclude(class_name__exact='')
            .values_list('class_name', flat=True)
        ))

        level_class_map[level] = sorted(level_classes_names, key=custom_sort)

    return Response({
        'levels': levels,
        'classes': classes,
        'level_class_map': level_class_map
    })

from .resources import StudentResource
from .import_utils import parse_student_file
import tablib
import tempfile
import os

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_update_file(request):
    """
    Handles file upload for bulk update/import of students.
    Supports Ministry formats (HTML, XLS, XLSX) via parse_student_file.
    """
    if not hasattr(request.user, 'profile') or not request.user.profile.has_perm('import_data'):
         return Response({'error': 'Unauthorized'}, status=403)

    if 'file' not in request.FILES:
        return Response({'error': 'No file uploaded'}, status=400)

    file_obj = request.FILES['file']

    # Save to temp file to allow processing by external libraries (xlrd, openpyxl, bs4)
    suffix = os.path.splitext(file_obj.name)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        for chunk in file_obj.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        # 1. Parse using custom logic (handles Ministry format & normalization)
        data_list = parse_student_file(tmp_path)

        if not data_list:
             return Response({'error': 'Could not parse file or no data found. Ensure format is correct.'}, status=400)

        # 2. Convert to Dataset for Import-Export
        dataset = tablib.Dataset()

        # Define headers matching StudentResource fields
        headers = [
            'student_id_number', 'last_name', 'first_name', 'gender',
            'date_of_birth', 'place_of_birth', 'academic_year',
            'class_name', 'attendance_system', 'enrollment_number',
            'enrollment_date'
        ]
        dataset.headers = headers

        for item in data_list:
            # Map parse result to resource structure
            row = [
                item.get('student_id_number'),
                item.get('last_name'),
                item.get('first_name'),
                item.get('gender'),
                item.get('date_of_birth'),
                item.get('place_of_birth'),
                item.get('academic_year'),
                item.get('class_name'),
                item.get('attendance_system', 'نصف داخلي'),
                item.get('enrollment_number'),
                item.get('enrollment_date')
            ]
            dataset.append(row)

        # 3. Execute Import (Update or Create)
        resource = StudentResource()
        result = resource.import_data(dataset, dry_run=False, raise_errors=True)

        return Response({
            'message': 'Import successful',
            'stats': {
                'total_processed': result.total_rows,
                'new_records': result.totals.get('new', 0),
                'updated_records': result.totals.get('update', 0),
                'skipped': result.totals.get('skip', 0),
                'errors': result.totals.get('error', 0)
            }
        })

    except Exception as e:
        logger.error(f"Import Error: {e}")
        return Response({'error': f"Import Failed: {str(e)}"}, status=500)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

class SystemMessageViewSet(viewsets.ModelViewSet):
    serializer_class = SystemMessageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = SystemMessage.objects.filter(active=True)
        # If recipient is defined, it only applies to that user OR the user is director
        from django.db.models import Q
        return qs.filter(Q(recipient__isnull=True) | Q(recipient=self.request.user))

    def perform_create(self, serializer):
        if hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'director':
            # Ensure empty string for recipient is saved as None so it applies to everyone
            recipient = self.request.data.get('recipient')
            if recipient == "" or recipient == "null":
                serializer.save(recipient=None)
            else:
                serializer.save()
        else:
            raise PermissionError("Only Director can create messages")
